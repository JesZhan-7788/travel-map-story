from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "local-imports"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "\n").strip()


def workbook_snapshot(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows: list[dict[str, Any]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [cell_text(value) for value in row]
            if any(values):
                rows.append({"row": row_index, "values": values})

        sheets.append(
            {
                "name": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "rows": rows,
            }
        )

    workbook.close()
    return {"source": path.name, "sheets": sheets}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for workbook in sorted(ROOT.glob("*.xlsx")):
        snapshot = workbook_snapshot(workbook)
        output = OUTPUT_DIR / f"{workbook.stem}.raw.json"
        output.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"wrote {output}")

    print("Raw imports are local-only. Review and manually promote sanitized content into src/data/trips.ts.")


if __name__ == "__main__":
    main()
